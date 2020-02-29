from dataprep import prep_mnist, add_noise
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from openpyxl import Workbook, load_workbook
from copy import deepcopy
from time import time


def bulk_train_noise(train_data, test_data, workbook, noise_range=(0, 20, 4), estimator_range=(1,11,1), n_jobs=8):
    """Bulk trains classifiers over a range of noise"""
    # Prepare our workbook
    wb = Workbook()

    # Pull our options from our training labels
    opts = np.unique(train_data[1]).tolist()

    # Iterate over a noise range and generate data
    for n in range(noise_range[0], noise_range[1], noise_range[2]):
        # Worksheet management
        ws = wb.create_sheet(title='noise-' + str(n))
        # First we need to prepare our new data.
        x_train = train_data[0]
        y_train = deepcopy(train_data[1])
        if (n != 0):
            (y_train, rnoise) = add_noise(y_train, opts, n)
        print("Beginning training on " + str(n) + "% noise")
        bulk_train_classifiers(train_data=(x_train, y_train), test_data=test_data, worksheet=ws, estimator_range=estimator_range, noise=n, n_jobs=n_jobs)
        print("Finished training on " + str(n) + "% noise")
    wb.save(filename=workbook)
    wb.close()




def bulk_train_classifiers(train_data, test_data, worksheet, estimator_range=(1, 2, 1), noise=0, n_jobs=8):
    """Bulk trains classifiers over a range of estimators"""
    worksheet.append(["n_estimators", "noise", "accuracy"])
    p=10
    for i in range(estimator_range[0], estimator_range[1], estimator_range[2]):
        train_classifier(train_data=train_data, test_data=test_data, worksheet=worksheet, n_estimators=i, n_jobs=n_jobs, noise=noise)
        if(i/estimator_range[1]>=p/100.0):
            p+=10
            print("Noise(" + str(noise) + ") progress - " + str(i / estimator_range[1] * 100) + "% - Estimator Number - " + str(i))


# Trains a single classifier and dumps it into an excel workbook
def train_classifier(train_data, test_data, worksheet, n_estimators=1, n_jobs=8, noise=0):
    # Creates and trains our classifier
    clf = RandomForestClassifier(n_estimators=n_estimators, n_jobs=n_jobs)
    clf = clf.fit(train_data[0], train_data[1])
    accuracy = clf.score(test_data[0], test_data[1])
    # Loads and appends our data to the current workbook
    worksheet.append([n_estimators, noise, accuracy])


if __name__ == "__main__":
    (x_train, y_train), (x_test, y_test) = prep_mnist()
    bulk_train_noise((x_train, y_train), (x_test, y_test), noise_range=(0, 20, 2), n_jobs=24, estimator_range=(1, 1001, 2), workbook="classifiers.xlsx")
