from dataprep import prep_mnist, add_noise
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from openpyxl import Workbook, load_workbook
from copy import deepcopy
from time import time


def train_classifier(train_data, test_data, og_labels, max_leaves, worksheet, n_estimators=1, n_jobs=8):
    """Trains a single classifier and dumps it into an excel workbook"""
    # Creates and trains our classifier
    clf = RandomForestClassifier(n_estimators=n_estimators, n_jobs=n_jobs, max_leaf_nodes=max_leaves)
    clf = clf.fit(train_data[0], train_data[1])
    test_accuracy = clf.score(test_data[0], test_data[1])
    train_accuracy = clf.score(train_data[0], train_data[1])
    og_train_accuracy = clf.score(train_data[0], og_labels)
    # Loads and appends our data to the current workbook
    worksheet.append([n_estimators, max_leaves, (max_leaves * n_estimators), train_accuracy, og_train_accuracy, test_accuracy])

def bulk_train_classifiers(train_data, test_data, workbook_name="results.xlsx", leaf_range=(10, 1000, 2000), noise_range=(0, 5, 10, 20), estimator_range=(1, 10, 20), n_jobs=8):
    # Prepare our workbook
    wb = Workbook()
    # Our timer
    start_time = time()
    # Pull our options from our training labels
    opts = np.unique(train_data[1]).tolist()

    for noise in noise_range:
        worksheet = wb.create_sheet(title='noise-' + str(noise))
        worksheet.append(["n_estimators", "leaves", "complexity", "train_accuracy", "train og labels", "test_accuracy"])
        x_train = train_data[0]
        y_train = deepcopy(train_data[1])
        if (noise != 0):
            (y_train, rnoise) = add_noise(y_train, opts, noise)
        for n_leaves in leaf_range:
            for estimators in estimator_range:
                train_classifier(train_data=(x_train, y_train), test_data=test_data, og_labels=train_data[1], worksheet=worksheet, n_estimators=estimators, n_jobs=n_jobs, max_leaves=n_leaves)

    wb.save(filename=workbook_name)
    wb.close()

if __name__ == "__main__":
    (x_train, y_train), (x_test, y_test) = prep_mnist()
    bulk_train_classifiers((x_train, y_train), (x_test, y_test), n_jobs=24, leaf_range=(2, 2000), estimator_range=range(5, 100, 5))
